# -*- coding: utf-8 -*-

import openpyxl


class t_Values:
    def __init__(self):
        self.values_dict = {}

    def get_values(self, t_id):
        wb = openpyxl.load_workbook('t_link.xlsx')
        sheet = wb.get_active_sheet()
        row_len = sheet.max_row
        for line in range(2, row_len+1):
            # a = sheet.cell(row=line, column=1).value
            b = sheet.cell(row=line, column=2).value
            c = sheet.cell(row=line, column=3).value
            e = sheet.cell(row=line, column=5).value
            f = sheet.cell(row=line, column=6).value
            case = [b, c, e, f]
            k = str(int(t_id)+line-2)
            self.values_dict[k] = case
        return self.values_dict
